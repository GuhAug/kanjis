import openpyxl
import json
import os

XLSX_PATH = r"C:\Users\GuhAu\Downloads\Kanji_New_Progressive_1-4.xlsx"
OUT_PATH  = r"C:\Users\GuhAu\kanjis\data\kanji.js"

LEVEL_META = [
    (1, "Iniciante",      "New Progressive 1"),
    (2, "Elementar",      "New Progressive 2"),
    (3, "Intermediário",  "New Progressive 3"),
    (4, "Avançado",       "New Progressive 4"),
]

CHAPTER_MAP = {
    "漢字１": 1, "漢字２": 2, "漢字３": 3, "漢字４": 4,
    "漢字５": 5, "漢字６": 6, "漢字７": 7, "漢字８": 8,
    "漢字1":  1, "漢字2":  2, "漢字3":  3, "漢字4":  4,
    "漢字5":  5, "漢字6":  6, "漢字7":  7, "漢字8":  8,
}

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("—", "-", "", "ー") else s

wb = openpyxl.load_workbook(XLSX_PATH)
records = []
uid = 0

for level, level_name, sheet_name in LEVEL_META:
    if sheet_name not in wb.sheetnames:
        print(f"Warning: sheet '{sheet_name}' not found, skipping.")
        continue
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        kanji  = clean(row[0])
        kun    = clean(row[1])
        on_r   = clean(row[2])
        ch_raw = clean(row[3])
        pt     = clean(row[4])
        kun_ex = clean(row[5])
        kun_tr = clean(row[6])
        on_ex  = clean(row[7])
        on_tr  = clean(row[8]) if len(row) > 8 else None

        if not kanji:
            continue

        ch_num = CHAPTER_MAP.get(ch_raw or "", 0)

        records.append({
            "id":      uid,
            "level":   level,
            "lname":   level_name,
            "chapter": ch_num,
            "k":       kanji,
            "kun":     kun,
            "on":      on_r,
            "pt":      pt,
            "kunEx":   kun_ex,
            "kunTr":   kun_tr,
            "onEx":    on_ex,
            "onTr":    on_tr,
        })
        uid += 1

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
json_str = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
with open(OUT_PATH, "w", encoding="utf-8") as f:
    f.write(f"window.KANJI_DATA={json_str};\n")

print(f"OK: {len(records)} kanji extraídos para {OUT_PATH}")
